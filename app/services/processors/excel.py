"""Excel document processors with optimized chunking"""

from datetime import datetime
from typing import List, Dict, Any

from langchain_core.documents import Document
import openpyxl
import xlrd

from .base import DocumentProcessor
from app.core.config import get_settings


class ExcelProcessor(DocumentProcessor):
    """
    Excel (.xlsx) 处理：Row-to-Text策略
    将每行数据结合表头转换为语义文本，保留结构信息
    """

    @staticmethod
    def process(
        file_path: str,
        metadata: Dict[str, Any],
        chunk_size: int = 1000,
        chunk_overlap: int = 200,
    ) -> List[Document]:
        """
        Process XLSX file using Row-to-Text strategy

        Converts each row into semantic text combined with headers.
        This preserves the structured nature of Excel data while
        enabling semantic search.

        Args:
            file_path: Path to XLSX file
            metadata: Metadata to attach to documents
            chunk_size: Not used for Excel (each row is a document)
            chunk_overlap: Not used for Excel

        Returns:
            List of Document objects, one per row
        """
        settings = get_settings()
        rows_per_chunk = settings.excel_rows_per_chunk
        key_fields = set(settings.excel_key_fields)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        documents = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                continue

            # 第一行作为表头
            headers = [str(h).strip() if h else "" for h in rows[0]]

            # Filter headers: only keep key fields
            key_indices = [
                i
                for i, h in enumerate(headers)
                if h and any(key_field in h for key_field in key_fields)
            ]

            if not key_indices:
                # If no key fields found, use all fields (fallback)
                key_indices = list(range(len(headers)))

            # 批量处理数据行：每 N 行合并为一个 chunk
            data_rows = rows[1:]
            for batch_start in range(0, len(data_rows), rows_per_chunk):
                batch_rows = data_rows[batch_start : batch_start + rows_per_chunk]
                batch_texts = []

                for row_offset, row in enumerate(batch_rows):
                    row_idx = (
                        batch_start + row_offset + 2
                    )  # +2 for header and 1-based index
                    row_data = []

                    # Only process key fields
                    for idx in key_indices:
                        if idx < len(row):
                            header = headers[idx]
                            value = row[idx]

                            if header and value is not None:
                                # 处理日期时间类型
                                value_str = (
                                    value.strftime("%Y-%m-%d %H:%M:%S")
                                    if isinstance(value, datetime)
                                    else str(value).strip()
                                )
                                if value_str:
                                    row_data.append(f"{header}: {value_str}")

                    if row_data:
                        batch_texts.append(f"第{row_idx}行 - {' | '.join(row_data)}")

                if batch_texts:
                    # 合并多行为一个文档
                    combined_text = "\n".join(batch_texts)
                    row_range_start = batch_start + 2
                    row_range_end = min(
                        batch_start + rows_per_chunk + 1, len(data_rows) + 1
                    )

                    doc_metadata = {
                        **metadata,
                        "sheet_name": sheet_name,
                        "row_range": f"{row_range_start}-{row_range_end}",
                        "content_type": "excel_batch",
                        "source_type": "excel",
                    }
                    documents.append(
                        Document(page_content=combined_text, metadata=doc_metadata)
                    )

        wb.close()
        return documents


class LegacyExcelProcessor(DocumentProcessor):
    """旧版Excel (.xls) 处理：Row-to-Text策略"""

    @staticmethod
    def process(
        file_path: str,
        metadata: Dict[str, Any],
        chunk_size: int = 1000,
        chunk_overlap: int = 200,
    ) -> List[Document]:
        """
        Process legacy XLS file using Row-to-Text strategy

        Args:
            file_path: Path to XLS file
            metadata: Metadata to attach to documents
            chunk_size: Not used for Excel (each row is a document)
            chunk_overlap: Not used for Excel

        Returns:
            List of Document objects, one per row
        """
        settings = get_settings()
        rows_per_chunk = settings.excel_rows_per_chunk
        key_fields = set(settings.excel_key_fields)

        workbook = xlrd.open_workbook(file_path)
        documents = []

        for sheet in workbook.sheets():
            sheet_name = sheet.name

            if sheet.nrows == 0:
                continue

            # 第一行作为表头
            headers = [
                (
                    str(sheet.cell_value(0, col)).strip()
                    if sheet.cell_value(0, col)
                    else ""
                )
                for col in range(sheet.ncols)
            ]

            # Filter headers: only keep key fields
            key_indices = [
                i
                for i, h in enumerate(headers)
                if h and any(key_field in h for key_field in key_fields)
            ]

            if not key_indices:
                # If no key fields found, use all fields (fallback)
                key_indices = list(range(len(headers)))

            # 批量处理数据行：每 N 行合并为一个 chunk
            total_rows = sheet.nrows - 1  # Excluding header
            for batch_start in range(1, sheet.nrows, rows_per_chunk):
                batch_end = min(batch_start + rows_per_chunk, sheet.nrows)
                batch_texts = []

                for row_idx in range(batch_start, batch_end):
                    row_data = []

                    # Only process key fields
                    for col_idx in key_indices:
                        if col_idx < sheet.ncols:
                            header = headers[col_idx]
                            cell_value = sheet.cell_value(row_idx, col_idx)

                            if header and cell_value:
                                # 处理日期类型
                                if (
                                    sheet.cell_type(row_idx, col_idx)
                                    == xlrd.XL_CELL_DATE
                                ):
                                    date_tuple = xlrd.xldate_as_tuple(
                                        cell_value, workbook.datemode
                                    )
                                    value_str = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                                else:
                                    value_str = str(cell_value).strip()

                                if value_str:
                                    row_data.append(f"{header}: {value_str}")

                    if row_data:
                        batch_texts.append(
                            f"第{row_idx + 1}行 - {' | '.join(row_data)}"
                        )

                if batch_texts:
                    # 合并多行为一个文档
                    combined_text = "\n".join(batch_texts)

                    doc_metadata = {
                        **metadata,
                        "sheet_name": sheet_name,
                        "row_range": f"{batch_start + 1}-{batch_end}",
                        "content_type": "excel_batch",
                        "source_type": "excel",
                    }
                    documents.append(
                        Document(page_content=combined_text, metadata=doc_metadata)
                    )

        return documents
